#!/usr/bin/env python3
"""発行コード.xlsxを billing_code へ安全に取り込む。

既定では照合レポートだけを出力する。--apply は service role key と
--confirm-apply の両方を指定した場合だけ、新規の請求コードテーブルへ書き込む。
"""
from __future__ import annotations

import argparse
import json
import os
import re
import unicodedata
from pathlib import Path
from typing import Any
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from openpyxl import load_workbook


def text(value: Any) -> str:
    return unicodedata.normalize("NFKC", str(value or "")).replace("\u3000", " ").strip()


def read_codes(path: Path) -> dict[str, list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    result: dict[str, list[dict[str, Any]]] = {}
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        header_index = next((index for index, row in enumerate(rows) if any(text(cell).startswith("発行先コード") for cell in row)), None)
        if header_index is None:
            continue
        headers = [text(cell) for cell in rows[header_index]]
        code_index = next(index for index, value in enumerate(headers) if value.startswith("発行先コード"))
        name_index = next((index for index, value in enumerate(headers) if "請求先会社" in value or "契約名義" in value), None)
        notes_index = next((index for index, value in enumerate(headers) if value == "備考"), None)
        if name_index is None:
            continue
        items: list[dict[str, Any]] = []
        for row_number, row in enumerate(rows[header_index + 2 :], start=header_index + 3):
            code, name = text(row[code_index] if code_index < len(row) else None), text(row[name_index] if name_index < len(row) else None)
            if not code or not name or not re.fullmatch(r"[0-9]+", code):
                continue
            notes = text(row[notes_index] if notes_index is not None and notes_index < len(row) else None)
            items.append({"issue_code": code, "recipient_name": name, "source_sheet_name": sheet.title, "source_row_number": row_number, "notes": notes or None, "is_active": not bool(re.search(r"解約|使用不可|空き番号", name + notes))})
        result[sheet.title] = items
    return result


class Rest:
    def __init__(self, url: str, key: str):
        self.base_url = url.rstrip("/") + "/rest/v1"
        self.headers = {"apikey": key, "Authorization": f"Bearer {key}", "Content-Type": "application/json"}

    def request(self, method: str, table: str, query: dict[str, str] | None = None, body: Any = None, prefer: str | None = None) -> list[dict[str, Any]]:
        headers = dict(self.headers)
        if prefer:
            headers["Prefer"] = prefer
        url = f"{self.base_url}/{table}" + (f"?{urlencode(query)}" if query else "")
        request = Request(url, data=json.dumps(body, ensure_ascii=False).encode("utf-8") if body is not None else None, headers=headers, method=method)
        with urlopen(request) as response:
            payload = response.read().decode("utf-8")
        return json.loads(payload) if payload else []

    def property_for_sheet(self, sheet_name: str) -> str | None:
        if not sheet_name.isdigit():
            return None
        rows = self.request("GET", "asset_master", {
            "select": "asset_id", "asset_code": f"eq.{sheet_name}", "limit": "2",
        })
        return rows[0].get("asset_id") if len(rows) == 1 else None

    def tenant_ids_for_code(self, property_id: str, code: str) -> list[str]:
        rows = self.request("GET", "lease_contract_unit", {
            "select": "contract:lease_contract!inner(tenant_id,tenant:tenant_master!inner(external_tenant_code)),unit:unit_master!inner(property_id)",
            "unit.property_id": f"eq.{property_id}",
            "contract.tenant.external_tenant_code": f"eq.{code}",
            "limit": "1000",
        })
        tenant_ids: set[str] = set()
        for row in rows:
            contract = row.get("contract") or {}
            if isinstance(contract, list):
                contract = contract[0] if contract else {}
            tenant_id = contract.get("tenant_id")
            if tenant_id:
                tenant_ids.add(tenant_id)
        return sorted(tenant_ids)


def build_import_rows(by_sheet: dict[str, list[dict[str, Any]]], client: Rest) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    prepared: list[dict[str, Any]] = []
    report: dict[str, Any] = {"sheets": {}, "summary": {"matched": 0, "unmatched": 0, "review_required": 0, "skipped": 0}, "examples": []}
    for sheet_name, rows in by_sheet.items():
        property_id = client.property_for_sheet(sheet_name)
        sheet_summary = {"rows": len(rows), "property_id": property_id, "matched": 0, "unmatched": 0, "review_required": 0, "skipped": 0}
        for row in rows:
            if not property_id:
                sheet_summary["skipped"] += 1
                report["summary"]["skipped"] += 1
                continue
            tenant_ids = client.tenant_ids_for_code(property_id, row["issue_code"])
            match_status = "matched" if len(tenant_ids) == 1 else "unmatched" if not tenant_ids else "review_required"
            prepared.append({
                **row,
                "property_id": property_id,
                "tenant_id": tenant_ids[0] if len(tenant_ids) == 1 else None,
                "match_status": match_status,
                "is_primary": len(tenant_ids) == 1,
                "is_active": len(tenant_ids) == 1 and row["is_active"],
            })
            sheet_summary[match_status] += 1
            report["summary"][match_status] += 1
            if match_status != "matched" and len(report["examples"]) < 30:
                report["examples"].append({"sheet": sheet_name, "row": row["source_row_number"], "issue_code": row["issue_code"], "recipient_name": row["recipient_name"], "status": match_status})
        report["sheets"][sheet_name] = sheet_summary
    return prepared, report


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--apply", action="store_true", help="Supabaseへ登録する")
    parser.add_argument("--confirm-apply", action="store_true", help="照合結果を確認済みとして実際の登録を許可する")
    parser.add_argument("--supabase-url", default=os.getenv("VITE_SUPABASE_URL"))
    parser.add_argument("--service-role-key", default=os.getenv("SUPABASE_SERVICE_ROLE_KEY"))
    args = parser.parse_args()
    by_sheet = read_codes(args.workbook)
    if not args.apply:
        print(json.dumps({"sheets": {sheet: {"rows": len(rows)} for sheet, rows in by_sheet.items()}, "message": "照合には --apply と接続情報が必要です。DBへの書き込みは行っていません。"}, ensure_ascii=False, indent=2))
        return
    if not args.confirm_apply:
        raise SystemExit("実際に登録するには --confirm-apply を指定してください。")
    if not args.supabase_url or not args.service_role_key:
        raise SystemExit("--apply には SUPABASE_URL と SUPABASE_SERVICE_ROLE_KEY が必要です。")
    client = Rest(args.supabase_url, args.service_role_key)
    payload, report = build_import_rows(by_sheet, client)
    if payload:
        client.request("POST", "billing_code", {"on_conflict": "property_id,issue_code"}, payload, "resolution=merge-duplicates")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
