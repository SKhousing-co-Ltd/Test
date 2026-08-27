#!/usr/bin/env python3
"""Import a current rent roll; run without --apply before writing to Supabase."""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import unicodedata
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from hashlib import sha256
from pathlib import Path
from typing import Any
from urllib.error import HTTPError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from openpyxl import load_workbook

SOURCE_SYSTEM = "rent_roll_xlsx"
SPECIAL_LAYOUT_SHEETS = {"神戸", "福島", "ORH", "中之島"}
PROPERTY_NAME_ALIASES: dict[str, str] = {
    "三共郡山ビル北館": "三共ビル郡山北館",
    "三共郡山ビル南館": "三共ビル郡山南館",
}
EMPTY_TENANT_VALUES = {"", "空室", "空室（倉庫）", "空"}
SHEET_UNIT_COLUMN_OVERRIDES = {"梅田": 4}


@dataclass
class ImportIssue:
    source_sheet_name: str
    source_row_number: int | None
    issue_type: str
    message: str
    source_payload: dict[str, Any] = field(default_factory=dict)


@dataclass
class RentRollRecord:
    source_sheet_name: str
    source_row_number: int
    property_name: str
    source_status: str
    wing_code: str | None
    floor_label: str | None
    unit_code: str
    unit_type: str
    tenant_code: str | None
    tenant_name: str | None
    area_sqm: float | None
    monthly_rent_amount: int | None
    monthly_common_charge_amount: int | None
    monthly_parking_amount: int | None
    other_monthly_amount: int | None
    deposit_amount: int | None
    security_deposit_amount: int | None
    key_money_amount: int | None
    renewal_fee_amount: int | None
    contract_start_date: str | None
    contract_end_date: str | None
    renewal_terms: str | None
    payment_terms: str | None
    source_unit_discriminator: str = ""
    source_contract_discriminator: str = ""


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return unicodedata.normalize("NFKC", str(value)).replace("\u3000", " ").strip()


def normalize_identifier(value: Any) -> str:
    return re.sub(r"\s+", "", normalize_text(value)).replace("～", "〜")


def normalize_tenant_name(value: Any) -> str:
    return re.sub(r"\s+", "", normalize_text(value)).lower()


def primary_tenant_code(value: str | None) -> str | None:
    codes = tenant_codes(value)
    return codes[0] if codes else None


def tenant_codes(value: str | None) -> list[str]:
    if not value:
        return []
    return list(dict.fromkeys(
        normalize_identifier(part) for part in re.split(r"[\n/]", value) if normalize_identifier(part)
    ))


def contract_source_key(record: RentRollRecord) -> str:
    """Keep a contract stable when rows move within a source worksheet."""
    tenant_identity = primary_tenant_code(record.tenant_code) or normalize_tenant_name(record.tenant_name)
    raw = "|".join((record.property_name, record.wing_code or "", record.floor_label or "", record.unit_code, tenant_identity))
    base_key = f"rr:v2:{sha256(raw.encode('utf-8')).hexdigest()}"
    if not record.source_contract_discriminator:
        return base_key
    instance_raw = f"{base_key}|{record.source_contract_discriminator}"
    return f"rr:v3:{sha256(instance_raw.encode('utf-8')).hexdigest()}"


def import_notes(record: RentRollRecord) -> str | None:
    notes: list[str] = []
    if record.source_status == "解約予定":
        notes.append("レントロール上の状態: 解約予定（契約状態は active のまま保持）")
    if record.tenant_code and primary_tenant_code(record.tenant_code) != normalize_identifier(record.tenant_code):
        notes.append(f"元テナントコード: {record.tenant_code}")
    if record.source_unit_discriminator:
        notes.append(f"暫定区画識別子: {record.source_unit_discriminator}")
    return "\n".join(notes) or None


def as_number(value: Any) -> int | float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return value
    text = normalize_text(value).replace(",", "").replace("¥", "").replace("￥", "")
    return float(text) if re.fullmatch(r"-?\d+(?:\.\d+)?", text) else None


def as_date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = normalize_text(value)
    for pattern in (r"(\d{4})[/-](\d{1,2})[/-](\d{1,2})", r"(\d{4})年(\d{1,2})月(\d{1,2})日"):
        match = re.fullmatch(pattern, text)
        if match:
            return date(*map(int, match.groups())).isoformat()
    return None


def infer_unit_type(unit_code: str) -> str:
    if any(token in unit_code for token in ("駐輪", "自転車", "バイク")):
        return "bicycle_parking"
    if any(token in unit_code for token in ("駐車", "車庫", "パーキング")):
        return "parking"
    if any(token in unit_code for token in ("看板", "サイン")):
        return "signage"
    if any(token in unit_code for token in ("アンテナ", "基地局")):
        return "antenna"
    if "倉庫" in unit_code:
        return "warehouse"
    if any(token in unit_code for token in ("住居", "住宅", "居室")):
        return "residential"
    if "ATM" in unit_code or "機械" in unit_code:
        return "other"
    return "office"


def get_property_name(sheet: Any) -> str | None:
    title = normalize_text(sheet.cell(1, 3).value)
    match = re.search(r"【(.+?)】", title)
    if match:
        return PROPERTY_NAME_ALIASES.get(match.group(1), match.group(1))
    return PROPERTY_NAME_ALIASES.get(sheet.title)


def find_column(headers: dict[int, str], candidates: tuple[str, ...]) -> int | None:
    for column, header in headers.items():
        compact = re.sub(r"\s+", "", header)
        if any(candidate in compact for candidate in candidates):
            return column
    return None


def find_columns(headers: dict[int, str], candidates: tuple[str, ...]) -> list[int]:
    return [
        column
        for column, header in headers.items()
        if any(candidate in re.sub(r"\s+", "", header) for candidate in candidates)
    ]


def detect_columns(sheet: Any) -> tuple[dict[str, Any], int]:
    headers: dict[int, str] = {}
    header_row = 0
    for column in range(1, sheet.max_column + 1):
        # レントロールの見出しは先頭 3 行まで。データ行まで走査すると、
        # 「入居中」などの値に含まれる「室」を区画列と誤認してしまう。
        values = [normalize_text(sheet.cell(row, column).value) for row in range(1, min(sheet.max_row, 3) + 1)]
        headers[column] = " ".join(value for value in values if value)
        if any(value in {"室", "テナント名", "コード", "棟"} for value in values):
            header_row = max(header_row, max(index + 1 for index, value in enumerate(values) if value))
    columns = {
        "wing": find_column(headers, ("棟",)), "floor": find_column(headers, ("階-室", "階")),
        "unit": find_column(headers, ("階-室", "室")), "tenant_code": find_column(headers, ("コード",)),
        "tenant_name": find_column(headers, ("テナント名",)), "area": find_column(headers, ("面積",)),
        "rent": find_column(headers, ("賃料",)), "common_charge": find_column(headers, ("共益費",)),
        "parking_fee": find_column(headers, ("駐車場代", "駐車料")),
        "other_fee_columns": find_columns(headers, ("TVアンテナ", "アンテナ料", "看板代", "清掃代")),
        "deposit": find_column(headers, ("敷金",)), "security_deposit": find_column(headers, ("保証金",)),
        "key_money": find_column(headers, ("礼金",)), "renewal_fee": find_column(headers, ("更新料",)),
        "contract_start": find_column(headers, ("契約開始", "開始日", "始期", "契約日")),
        "contract_end": find_column(headers, ("契約終了", "終了日", "満了日", "終期", "更新日")),
        "renewal_terms": find_column(headers, ("更新条件", "更新周期")), "payment_terms": find_column(headers, ("支払条件", "支払期日")),
    }
    # 新大阪・横浜などは「室」の見出しが空欄だが、階の右隣が区画列である。
    if columns["unit"] is None and columns["floor"] and columns["tenant_name"] and columns["area"]:
        columns["unit"] = columns["floor"] + 1
    if sheet.title in SHEET_UNIT_COLUMN_OVERRIDES:
        columns["unit"] = SHEET_UNIT_COLUMN_OVERRIDES[sheet.title]
    return columns, header_row


def cell(row: tuple[Any, ...], column: int | None) -> Any:
    return row[column - 1] if column and column <= len(row) else None


def sum_columns(row: tuple[Any, ...], columns: list[int]) -> int | float | None:
    values = [as_number(cell(row, column)) for column in columns]
    return sum(value or 0 for value in values) if any(value is not None for value in values) else None


def read_workbook(path: Path, selected_sheet_names: set[str] | None = None) -> tuple[list[RentRollRecord], list[ImportIssue]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    records: list[RentRollRecord] = []
    issues: list[ImportIssue] = []
    found_sheet_names: set[str] = set()
    for sheet in workbook.worksheets:
        if selected_sheet_names is not None and sheet.title not in selected_sheet_names:
            continue
        found_sheet_names.add(sheet.title)
        if sheet.title in SPECIAL_LAYOUT_SHEETS:
            issues.append(ImportIssue(sheet.title, None, "layout_not_supported", "棟・住居用の専用レイアウトのため、設定を追加するまで自動取込の対象外です。"))
            continue
        property_name = get_property_name(sheet)
        if not property_name:
            issues.append(ImportIssue(sheet.title, None, "property_not_detected", "タイトルから物件名を取得できません。PROPERTY_NAME_ALIASES に対応を追加してください。"))
            continue
        columns, header_row = detect_columns(sheet)
        if not header_row or any(columns[name] is None for name in ("unit", "tenant_name", "area")):
            issues.append(ImportIssue(sheet.title, None, "layout_not_supported", "必須列（室・テナント名・面積）を特定できません。", {"columns": columns}))
            continue
        inherited_floor: str | None = None
        inherited_wing: str | None = None
        for row_number, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            unit_code = normalize_identifier(cell(row, columns["unit"]))
            if not unit_code or "合計" in unit_code:
                continue
            floor = normalize_text(cell(row, columns["floor"]))
            wing = normalize_text(cell(row, columns["wing"]))
            inherited_floor = floor or inherited_floor
            inherited_wing = wing or inherited_wing
            tenant_name = normalize_text(cell(row, columns["tenant_name"])) or None
            tenant_code = normalize_text(cell(row, columns["tenant_code"])) or None
            source_status = normalize_text(row[0])
            payload = {"floor": inherited_floor, "unit": unit_code, "tenant_code": tenant_code, "tenant_name": tenant_name}
            if any(marker in unit_code for marker in ("〜", "~", "・")):
                issues.append(ImportIssue(sheet.title, row_number, "combined_unit", "結合区画として暫定登録します。", payload))
            if tenant_code and re.search(r"[\n/]", tenant_code):
                issues.append(ImportIssue(sheet.title, row_number, "multiple_tenant_codes", "全テナントコードを保存しました。請求項目ごとのコード割当を確認してください。", payload))
            records.append(RentRollRecord(
                sheet.title, row_number, property_name, source_status, inherited_wing, inherited_floor, unit_code, infer_unit_type(unit_code),
                tenant_code, tenant_name, as_number(cell(row, columns["area"])), as_number(cell(row, columns["rent"])),
                as_number(cell(row, columns["common_charge"])), as_number(cell(row, columns["parking_fee"])),
                sum_columns(row, columns["other_fee_columns"]), as_number(cell(row, columns["deposit"])),
                as_number(cell(row, columns["security_deposit"])), as_number(cell(row, columns["key_money"])),
                as_number(cell(row, columns["renewal_fee"])), as_date(cell(row, columns["contract_start"])),
                as_date(cell(row, columns["contract_end"])), normalize_text(cell(row, columns["renewal_terms"])) or None,
                normalize_text(cell(row, columns["payment_terms"])) or None))
    for missing_sheet_name in sorted((selected_sheet_names or set()) - found_sheet_names):
        issues.append(ImportIssue(missing_sheet_name, None, "sheet_not_found", "指定されたシートがワークブックにありません。"))

    tenant_names_by_code: dict[tuple[str, str], str] = {}
    for record in records:
        tenant_code = primary_tenant_code(record.tenant_code)
        if tenant_code and record.tenant_name:
            tenant_names_by_code.setdefault((record.source_sheet_name, tenant_code), record.tenant_name)
    for record in records:
        tenant_code = primary_tenant_code(record.tenant_code)
        if tenant_code and not record.tenant_name:
            record.tenant_name = tenant_names_by_code.get((record.source_sheet_name, tenant_code))

    latest_start_by_code: dict[tuple[str, str], str] = {}
    for record in records:
        tenant_code = primary_tenant_code(record.tenant_code)
        if not tenant_code:
            continue
        key = (record.source_sheet_name, tenant_code)
        if record.contract_start_date:
            latest_start_by_code[key] = record.contract_start_date
        elif key in latest_start_by_code:
            record.contract_start_date = latest_start_by_code[key]

    grouped: dict[tuple[str, str | None, str | None, str], list[RentRollRecord]] = {}
    for record in records:
        grouped.setdefault((record.property_name, record.wing_code, record.floor_label, record.unit_code), []).append(record)
    for group in grouped.values():
        if len(group) == 1:
            continue
        discriminators: dict[str, int] = {}
        for record in group:
            base = primary_tenant_code(record.tenant_code) or normalize_tenant_name(record.tenant_name) or f"row-{record.source_row_number}"
            discriminators[base] = discriminators.get(base, 0) + 1
            record.source_unit_discriminator = base if discriminators[base] == 1 else f"{base}-row-{record.source_row_number}"
            issues.append(ImportIssue(record.source_sheet_name, record.source_row_number, "temporary_unit_discriminator", "同一階・同一区画名が複数あるため、暫定識別子を付けて登録します。", {"floor": record.floor_label, "unit": record.unit_code, "discriminator": record.source_unit_discriminator, "tenant_code": record.tenant_code, "tenant_name": record.tenant_name}))

    contract_groups: dict[tuple[str, str], list[RentRollRecord]] = {}
    for record in records:
        tenant_identity = primary_tenant_code(record.tenant_code) or normalize_tenant_name(record.tenant_name)
        identity = "|".join((record.property_name, record.wing_code or "", record.floor_label or "", record.unit_code, tenant_identity))
        contract_groups.setdefault((record.source_sheet_name, identity), []).append(record)
    for group in contract_groups.values():
        start_dates = {record.contract_start_date or "" for record in group}
        if len(start_dates) <= 1:
            continue
        for record in group:
            record.source_contract_discriminator = record.contract_start_date or f"row-{record.source_row_number}"
    return records, issues


class SupabaseRest:
    def __init__(self, url: str, service_role_key: str):
        self.url = url.rstrip("/") + "/rest/v1"
        self.headers = {"apikey": service_role_key, "Authorization": f"Bearer {service_role_key}", "Content-Type": "application/json"}

    def request(self, method: str, table: str, *, query: dict[str, str] | None = None, body: Any = None, prefer: str | None = None) -> Any:
        url = f"{self.url}/{table}" + (("?" + urlencode(query)) if query else "")
        headers = dict(self.headers)
        if prefer:
            headers["Prefer"] = prefer
        request = Request(url, data=json.dumps(body, ensure_ascii=False).encode("utf-8") if body is not None else None, headers=headers, method=method)
        try:
            with urlopen(request) as response:
                payload = response.read().decode("utf-8")
                return json.loads(payload) if payload else None
        except HTTPError as error:
            raise RuntimeError(f"{method} {table} failed: {error.read().decode('utf-8', 'replace')}") from error

    def one(self, table: str, query: dict[str, str]) -> dict[str, Any] | None:
        rows = self.request("GET", table, query={**query, "limit": "1"})
        return rows[0] if rows else None

    def many(self, table: str, query: dict[str, str]) -> list[dict[str, Any]]:
        return self.request("GET", table, query=query) or []


def ensure_tenant_billing_codes(client: SupabaseRest, tenant_id: str, source_codes: list[str]) -> None:
    if not source_codes:
        return
    existing_codes = client.many("tenant_billing_code", {
        "select": "tenant_billing_code_id,billing_code,is_primary",
        "tenant_id": f"eq.{tenant_id}",
    })
    existing_by_code = {row["billing_code"]: row for row in existing_codes}
    primary = next((row for row in existing_codes if row["is_primary"]), None)
    for index, billing_code in enumerate(source_codes):
        if billing_code in existing_by_code:
            continue
        row = client.request("POST", "tenant_billing_code", body={
            "tenant_id": tenant_id,
            "billing_code": billing_code,
            "is_primary": primary is None and index == 0,
            "is_active": True,
            "sort_order": len(existing_codes) + index,
        }, prefer="return=representation")[0]
        existing_by_code[billing_code] = row
        if row["is_primary"]:
            primary = row
    if primary is None:
        return
    client.request("PATCH", "tenant_master", query={"tenant_id": f"eq.{tenant_id}"}, body={
        "external_tenant_code": primary["billing_code"],
    })
    accounts = client.many("income_expense_account_master", {
        "select": "account_id",
        "income_expense_type": "eq.収入",
    })
    existing_assignments = client.many("tenant_billing_code_account", {
        "select": "account_id",
        "tenant_id": f"eq.{tenant_id}",
    })
    assigned = {row["account_id"] for row in existing_assignments}
    for account in accounts:
        if account["account_id"] in assigned:
            continue
        client.request("POST", "tenant_billing_code_account", body={
            "tenant_id": tenant_id,
            "account_id": account["account_id"],
            "tenant_billing_code_id": primary["tenant_billing_code_id"],
        })


def terminate_source_contracts_for_unit(client: SupabaseRest, unit_id: str) -> None:
    allocations = client.many("lease_contract_unit", {"select": "lease_contract_id,lease_contract!inner(lease_contract_id,source_system,contract_status)", "unit_id": f"eq.{unit_id}", "lease_contract.source_system": f"eq.{SOURCE_SYSTEM}", "lease_contract.contract_status": "eq.active"})
    for allocation in allocations:
        client.request("PATCH", "lease_contract", query={"lease_contract_id": f"eq.{allocation['lease_contract_id']}"}, body={"contract_status": "terminated"})


def deactivate_imported_units_for_property(client: SupabaseRest, property_id: str) -> None:
    units = client.many("unit_master", {"select": "unit_id,source_discriminator", "property_id": f"eq.{property_id}"})
    for unit in units:
        source_allocations = client.many("lease_contract_unit", {
            "select": "lease_contract_id,lease_contract!inner(source_system)",
            "unit_id": f"eq.{unit['unit_id']}",
            "lease_contract.source_system": f"eq.{SOURCE_SYSTEM}",
        })
        if unit.get("source_discriminator") or source_allocations:
            client.request("PATCH", "unit_master", query={"unit_id": f"eq.{unit['unit_id']}"}, body={"is_active": False})


def terminate_stale_source_contracts_for_properties(client: SupabaseRest, property_ids: set[str], active_source_keys: set[str]) -> None:
    terminated_contract_ids: set[str] = set()
    for property_id in property_ids:
        units = client.many("unit_master", {"select": "unit_id", "property_id": f"eq.{property_id}"})
        for unit in units:
            allocations = client.many("lease_contract_unit", {
                "select": "lease_contract_id,lease_contract!inner(source_record_key,source_system,contract_status)",
                "unit_id": f"eq.{unit['unit_id']}",
                "lease_contract.source_system": f"eq.{SOURCE_SYSTEM}",
                "lease_contract.contract_status": "eq.active",
            })
            for allocation in allocations:
                contract = allocation.get("lease_contract") or {}
                contract_id = allocation["lease_contract_id"]
                if contract_id in terminated_contract_ids or contract.get("source_record_key") in active_source_keys:
                    continue
                client.request("PATCH", "lease_contract", query={"lease_contract_id": f"eq.{contract_id}"}, body={"contract_status": "terminated"})
                terminated_contract_ids.add(contract_id)


def persist(records: list[RentRollRecord], issues: list[ImportIssue], source_file_name: str, client: SupabaseRest) -> tuple[int, list[ImportIssue]]:
    persisted = 0
    runtime_issues = list(issues)
    active_source_keys: set[str] = set()
    property_rows: dict[str, dict[str, Any] | None] = {}
    for property_name in {record.property_name for record in records}:
        property_rows[property_name] = client.one("asset_master", {"select": "asset_id", "asset_name": f"eq.{property_name}"})
    imported_property_ids = {row["asset_id"] for row in property_rows.values() if row}
    for property_id in imported_property_ids:
        deactivate_imported_units_for_property(client, property_id)
    for source_sheet_name in {record.source_sheet_name for record in records} | {issue.source_sheet_name for issue in issues}:
        client.request("DELETE", "rent_roll_import_issue", query={"source_file_name": f"eq.{source_file_name}", "source_sheet_name": f"eq.{source_sheet_name}"})
    for record in records:
        property_row = property_rows[record.property_name]
        if not property_row:
            runtime_issues.append(ImportIssue(record.source_sheet_name, record.source_row_number, "property_not_matched", "asset_master に一致する物件がありません。", asdict(record)))
            continue
        property_id = property_row["asset_id"]
        unit_query = {"select": "unit_id", "property_id": f"eq.{property_id}", "unit_code": f"eq.{record.unit_code}", "source_discriminator": f"eq.{record.source_unit_discriminator}"}
        unit_query["floor_label"] = f"eq.{record.floor_label}" if record.floor_label else "is.null"
        unit_payload: dict[str, Any] = {"property_id": property_id, "unit_code": record.unit_code, "floor_label": record.floor_label, "source_discriminator": record.source_unit_discriminator, "unit_type": record.unit_type, "rentable_area_sqm": record.area_sqm, "is_active": True}
        if record.wing_code:
            wing = client.one("building_wing_master", {"select": "building_wing_id", "property_id": f"eq.{property_id}", "wing_code": f"eq.{record.wing_code}"})
            if not wing:
                wing = client.request("POST", "building_wing_master", body={"property_id": property_id, "wing_code": record.wing_code, "wing_name": record.wing_code}, prefer="return=representation")[0]
            unit_query["building_wing_id"] = f"eq.{wing['building_wing_id']}"
            unit_payload["building_wing_id"] = wing["building_wing_id"]
        else:
            unit_query["building_wing_id"] = "is.null"
            unit_payload["building_wing_id"] = None
        unit = client.one("unit_master", unit_query)
        if unit:
            client.request("PATCH", "unit_master", query={"unit_id": f"eq.{unit['unit_id']}"}, body=unit_payload)
        else:
            unit = client.request("POST", "unit_master", body=unit_payload, prefer="return=representation")[0]
        if not record.tenant_name or record.tenant_name in EMPTY_TENANT_VALUES or record.source_status.startswith("空室"):
            terminate_source_contracts_for_unit(client, unit["unit_id"])
            persisted += 1
            continue
        normalized_name = normalize_tenant_name(record.tenant_name)
        tenant = client.one("tenant_master", {"select": "tenant_id,external_tenant_code", "normalized_tenant_name": f"eq.{normalized_name}"})
        tenant_payload = {"external_tenant_code": tenant.get("external_tenant_code") if tenant else primary_tenant_code(record.tenant_code), "tenant_name": record.tenant_name, "normalized_tenant_name": normalized_name}
        if tenant:
            client.request("PATCH", "tenant_master", query={"tenant_id": f"eq.{tenant['tenant_id']}"}, body=tenant_payload)
        else:
            tenant = client.request("POST", "tenant_master", body=tenant_payload, prefer="return=representation")[0]
        ensure_tenant_billing_codes(client, tenant["tenant_id"], tenant_codes(record.tenant_code))
        source_key = contract_source_key(record)
        contract = client.one("lease_contract", {"select": "lease_contract_id", "source_system": f"eq.{SOURCE_SYSTEM}", "source_record_key": f"eq.{source_key}"})
        contract_payload = {"tenant_id": tenant["tenant_id"], "contract_status": "active", "contract_start_date": record.contract_start_date, "contract_end_date": record.contract_end_date, "renewal_terms": record.renewal_terms, "payment_terms": record.payment_terms, "notes": import_notes(record), "source_system": SOURCE_SYSTEM, "source_record_key": source_key}
        if contract:
            client.request("PATCH", "lease_contract", query={"lease_contract_id": f"eq.{contract['lease_contract_id']}"}, body=contract_payload)
        else:
            contract = client.request("POST", "lease_contract", body=contract_payload, prefer="return=representation")[0]
        allocation = client.one("lease_contract_unit", {"select": "lease_contract_unit_id", "lease_contract_id": f"eq.{contract['lease_contract_id']}", "unit_id": f"eq.{unit['unit_id']}"})
        allocation_payload = {"lease_contract_id": contract["lease_contract_id"], "unit_id": unit["unit_id"], "leased_area_sqm": record.area_sqm, "monthly_rent_amount": record.monthly_rent_amount, "monthly_common_charge_amount": record.monthly_common_charge_amount, "deposit_amount": record.deposit_amount, "security_deposit_amount": record.security_deposit_amount, "key_money_amount": record.key_money_amount, "renewal_fee_amount": record.renewal_fee_amount}
        if allocation:
            client.request("PATCH", "lease_contract_unit", query={"lease_contract_unit_id": f"eq.{allocation['lease_contract_unit_id']}"}, body=allocation_payload)
        else:
            client.request("POST", "lease_contract_unit", body=allocation_payload)
        active_source_keys.add(source_key)
        persisted += 1
    terminate_stale_source_contracts_for_properties(client, imported_property_ids, active_source_keys)
    for issue in runtime_issues:
        client.request("POST", "rent_roll_import_issue", body={"source_file_name": source_file_name, **asdict(issue)})
    return persisted, runtime_issues


def write_sql_import(records: list[RentRollRecord], issues: list[ImportIssue], source_file_name: str, path: Path, *, wrap_transaction: bool = True) -> None:
    """Create one transactional import for the Supabase CLI when no service key is available."""
    rows = [{
        "source_sheet_name": record.source_sheet_name,
        "source_row_number": record.source_row_number,
        "property_name": record.property_name,
        "source_status": record.source_status,
        "wing_code": record.wing_code,
        "floor_label": record.floor_label,
        "unit_code": record.unit_code,
        "unit_type": record.unit_type,
        "source_discriminator": record.source_unit_discriminator,
        "tenant_code": primary_tenant_code(record.tenant_code),
        "tenant_codes": tenant_codes(record.tenant_code),
        "tenant_name": record.tenant_name,
        "normalized_tenant_name": normalize_tenant_name(record.tenant_name),
        "area_sqm": record.area_sqm,
        "monthly_rent_amount": record.monthly_rent_amount,
        "monthly_common_charge_amount": record.monthly_common_charge_amount,
        "deposit_amount": record.deposit_amount,
        "security_deposit_amount": record.security_deposit_amount,
        "key_money_amount": record.key_money_amount,
        "renewal_fee_amount": record.renewal_fee_amount,
        "contract_start_date": record.contract_start_date,
        "contract_end_date": record.contract_end_date,
        "renewal_terms": record.renewal_terms,
        "payment_terms": record.payment_terms,
        "notes": import_notes(record),
        "source_record_key": contract_source_key(record),
        "is_vacant": not record.tenant_name or record.tenant_name in EMPTY_TENANT_VALUES or record.source_status.startswith("空室"),
    } for record in records]
    issue_rows = [{
        "source_sheet_name": issue.source_sheet_name,
        "source_row_number": issue.source_row_number,
        "issue_type": issue.issue_type,
        "message": issue.message,
        "source_payload": issue.source_payload,
    } for issue in issues]
    source_file_sql = "'" + source_file_name.replace("'", "''") + "'"
    transaction_start = "begin;\n\n" if wrap_transaction else ""
    transaction_end = "\ncommit;\n" if wrap_transaction else "\n"
    sql = f'''{transaction_start}create temporary table rent_roll_stage (
  source_sheet_name text not null, source_row_number integer not null, property_name text not null,
  source_status text not null, wing_code text, floor_label text, unit_code text not null, unit_type text not null,
  source_discriminator text not null, tenant_code text, tenant_codes jsonb not null, tenant_name text, normalized_tenant_name text not null,
  area_sqm numeric, monthly_rent_amount numeric, monthly_common_charge_amount numeric, deposit_amount numeric,
  security_deposit_amount numeric, key_money_amount numeric, renewal_fee_amount numeric,
  contract_start_date date, contract_end_date date, renewal_terms text, payment_terms text, notes text,
  source_record_key text not null, is_vacant boolean not null
) on commit drop;

insert into rent_roll_stage
select * from jsonb_to_recordset($rentroll${json.dumps(rows, ensure_ascii=False)}$rentroll$::jsonb) as r(
  source_sheet_name text, source_row_number integer, property_name text, source_status text, wing_code text,
  floor_label text, unit_code text, unit_type text, source_discriminator text, tenant_code text, tenant_codes jsonb, tenant_name text,
  normalized_tenant_name text, area_sqm numeric, monthly_rent_amount numeric, monthly_common_charge_amount numeric,
  deposit_amount numeric, security_deposit_amount numeric, key_money_amount numeric, renewal_fee_amount numeric,
  contract_start_date date, contract_end_date date, renewal_terms text, payment_terms text, notes text,
  source_record_key text, is_vacant boolean
);

create temporary table rent_roll_active_source_key (source_record_key text primary key) on commit drop;

delete from public.rent_roll_import_issue issue
using (select distinct source_sheet_name from rent_roll_stage) source_sheet
where issue.source_file_name = {source_file_sql}
  and issue.source_sheet_name = source_sheet.source_sheet_name;

update public.unit_master unit
set is_active = false, updated_at = now()
where unit.property_id in (
  select asset.asset_id
  from public.asset_master asset
  join (select distinct property_name from rent_roll_stage) source_property
    on source_property.property_name = asset.asset_name
)
and (
  coalesce(unit.source_discriminator, '') <> ''
  or exists (
    select 1
    from public.lease_contract_unit contract_unit
    join public.lease_contract contract on contract.lease_contract_id = contract_unit.lease_contract_id
    where contract_unit.unit_id = unit.unit_id and contract.source_system = '{SOURCE_SYSTEM}'
  )
);

do $rentroll_import$
declare
  r rent_roll_stage%rowtype;
  v_property_id uuid;
  v_wing_id uuid;
  v_unit_id uuid;
  v_tenant_id uuid;
  v_billing_code text;
  v_primary_code_id uuid;
  v_contract_id uuid;
  v_allocation_id uuid;
begin
  for r in select * from rent_roll_stage loop
    select asset_id into v_property_id from public.asset_master where asset_name = r.property_name;
    if v_property_id is null then
      insert into public.rent_roll_import_issue (source_file_name, source_sheet_name, source_row_number, issue_type, message, source_payload)
      values ({source_file_sql}, r.source_sheet_name, r.source_row_number, 'property_not_matched', 'asset_master に一致する物件がありません。', to_jsonb(r));
      continue;
    end if;

    v_wing_id := null;
    if r.wing_code is not null and r.wing_code <> '' then
      select building_wing_id into v_wing_id from public.building_wing_master where property_id = v_property_id and wing_code = r.wing_code;
      if v_wing_id is null then
        insert into public.building_wing_master (property_id, wing_code, wing_name) values (v_property_id, r.wing_code, r.wing_code) returning building_wing_id into v_wing_id;
      end if;
    end if;

    select unit_id into v_unit_id from public.unit_master
      where property_id = v_property_id and building_wing_id is not distinct from v_wing_id
        and floor_label is not distinct from r.floor_label and unit_code = r.unit_code
        and source_discriminator = r.source_discriminator;
    if v_unit_id is null then
      insert into public.unit_master (property_id, building_wing_id, unit_code, floor_label, source_discriminator, unit_type, rentable_area_sqm)
      values (v_property_id, v_wing_id, r.unit_code, r.floor_label, r.source_discriminator, r.unit_type, r.area_sqm)
      returning unit_id into v_unit_id;
    else
      update public.unit_master set unit_type = r.unit_type, rentable_area_sqm = r.area_sqm, is_active = true, updated_at = now() where unit_id = v_unit_id;
    end if;

    if r.is_vacant then
      update public.lease_contract c set contract_status = 'terminated', updated_at = now()
      where c.contract_status = 'active' and c.source_system = '{SOURCE_SYSTEM}'
        and exists (select 1 from public.lease_contract_unit cu where cu.lease_contract_id = c.lease_contract_id and cu.unit_id = v_unit_id);
      continue;
    end if;

    select tenant_id into v_tenant_id from public.tenant_master where normalized_tenant_name = r.normalized_tenant_name;
    if v_tenant_id is null and r.tenant_code is not null then
      select tenant_id into v_tenant_id from public.tenant_master where external_tenant_code = r.tenant_code;
      if v_tenant_id is null then
        select tenant_id into v_tenant_id from public.tenant_billing_code where billing_code = r.tenant_code;
      end if;
    end if;
    if v_tenant_id is null then
      insert into public.tenant_master (external_tenant_code, tenant_name, normalized_tenant_name)
      values (r.tenant_code, r.tenant_name, r.normalized_tenant_name) returning tenant_id into v_tenant_id;
    else
      update public.tenant_master set external_tenant_code = coalesce(r.tenant_code, external_tenant_code), tenant_name = r.tenant_name, normalized_tenant_name = r.normalized_tenant_name, updated_at = now()
      where tenant_id = v_tenant_id;
    end if;

    select tenant_billing_code_id into v_primary_code_id
    from public.tenant_billing_code
    where tenant_id = v_tenant_id and is_primary;
    for v_billing_code in select jsonb_array_elements_text(r.tenant_codes) loop
      if not exists (
        select 1 from public.tenant_billing_code
        where tenant_id = v_tenant_id and billing_code = v_billing_code
      ) then
        if v_primary_code_id is null then
          insert into public.tenant_billing_code(tenant_id, billing_code, is_primary, is_active, sort_order)
          values (v_tenant_id, v_billing_code, true, true, 0)
          returning tenant_billing_code_id into v_primary_code_id;
        else
          insert into public.tenant_billing_code(tenant_id, billing_code, is_primary, is_active, sort_order)
          values (
            v_tenant_id, v_billing_code, false, true,
            coalesce((select max(sort_order) + 1 from public.tenant_billing_code where tenant_id = v_tenant_id), 0)
          );
        end if;
      end if;
    end loop;
    if v_primary_code_id is null then
      select tenant_billing_code_id into v_primary_code_id
      from public.tenant_billing_code where tenant_id = v_tenant_id and is_primary;
    end if;
    if v_primary_code_id is not null then
      insert into public.tenant_billing_code_account(tenant_id, account_id, tenant_billing_code_id)
      select v_tenant_id, account.account_id, v_primary_code_id
      from public.income_expense_account_master account
      where account.income_expense_type = '収入'
      on conflict (tenant_id, account_id) do nothing;
      update public.tenant_master tenant
      set external_tenant_code = code.billing_code, updated_at = now()
      from public.tenant_billing_code code
      where tenant.tenant_id = v_tenant_id
        and code.tenant_billing_code_id = v_primary_code_id
        and tenant.external_tenant_code is distinct from code.billing_code;
    end if;

    select lease_contract_id into v_contract_id from public.lease_contract
      where source_system = '{SOURCE_SYSTEM}' and source_record_key = r.source_record_key;
    if v_contract_id is null then
      insert into public.lease_contract (tenant_id, contract_status, contract_start_date, contract_end_date, renewal_terms, payment_terms, notes, source_system, source_record_key)
      values (v_tenant_id, 'active', r.contract_start_date, r.contract_end_date, r.renewal_terms, r.payment_terms, r.notes, '{SOURCE_SYSTEM}', r.source_record_key)
      returning lease_contract_id into v_contract_id;
    else
      update public.lease_contract set tenant_id = v_tenant_id, contract_status = 'active', contract_start_date = r.contract_start_date, contract_end_date = r.contract_end_date,
        renewal_terms = r.renewal_terms, payment_terms = r.payment_terms, notes = r.notes, updated_at = now() where lease_contract_id = v_contract_id;
    end if;

    select lease_contract_unit_id into v_allocation_id from public.lease_contract_unit where lease_contract_id = v_contract_id and unit_id = v_unit_id;
    if v_allocation_id is null then
      insert into public.lease_contract_unit (lease_contract_id, unit_id, leased_area_sqm, monthly_rent_amount, monthly_common_charge_amount, deposit_amount, security_deposit_amount, key_money_amount, renewal_fee_amount)
      values (v_contract_id, v_unit_id, r.area_sqm, r.monthly_rent_amount, r.monthly_common_charge_amount, r.deposit_amount, r.security_deposit_amount, r.key_money_amount, r.renewal_fee_amount);
    else
      update public.lease_contract_unit set leased_area_sqm = r.area_sqm, monthly_rent_amount = r.monthly_rent_amount, monthly_common_charge_amount = r.monthly_common_charge_amount,
        deposit_amount = r.deposit_amount, security_deposit_amount = r.security_deposit_amount, key_money_amount = r.key_money_amount, renewal_fee_amount = r.renewal_fee_amount, updated_at = now()
      where lease_contract_unit_id = v_allocation_id;
    end if;
    insert into rent_roll_active_source_key (source_record_key) values (r.source_record_key) on conflict do nothing;
  end loop;

  update public.lease_contract c set contract_status = 'terminated', updated_at = now()
  where c.source_system = '{SOURCE_SYSTEM}' and c.contract_status = 'active'
    and exists (
      select 1
      from public.lease_contract_unit contract_unit
      join public.unit_master unit on unit.unit_id = contract_unit.unit_id
      join public.asset_master asset on asset.asset_id = unit.property_id
      where contract_unit.lease_contract_id = c.lease_contract_id
        and asset.asset_name in (select distinct property_name from rent_roll_stage)
    )
    and not exists (select 1 from rent_roll_active_source_key k where k.source_record_key = c.source_record_key);
end;
$rentroll_import$;

insert into public.rent_roll_import_issue (source_file_name, source_sheet_name, source_row_number, issue_type, message, source_payload)
select {source_file_sql}, source_sheet_name, source_row_number, issue_type, message, source_payload
from jsonb_to_recordset($rentrollissues${json.dumps(issue_rows, ensure_ascii=False)}$rentrollissues$::jsonb) as i(
  source_sheet_name text, source_row_number integer, issue_type text, message text, source_payload jsonb
);

{transaction_end}'''
    path.write_text(sql, encoding="utf-8")


def write_reconciliation_sql(
    records: list[RentRollRecord],
    issues: list[ImportIssue],
    workbook: Path,
    as_of_date: str,
    path: Path,
) -> None:
    """Write comparison-only SQL. This never updates contract or rent-roll master data."""
    source_hash = sha256(workbook.read_bytes()).hexdigest()
    rows: list[dict[str, Any]] = []
    for record in records:
        monthly_values = (
            record.monthly_rent_amount,
            record.monthly_common_charge_amount,
            record.monthly_parking_amount,
            record.other_monthly_amount,
        )
        monthly_total = sum(value or 0 for value in monthly_values) if any(value is not None for value in monthly_values) else None
        rows.append({
            "source_sheet_name": record.source_sheet_name,
            "source_row_number": record.source_row_number,
            "property_name": record.property_name,
            "wing_code": record.wing_code,
            "floor_label": record.floor_label,
            "unit_code": record.unit_code,
            "unit_type": record.unit_type,
            "tenant_code": primary_tenant_code(record.tenant_code),
            "tenant_name": record.tenant_name,
            "source_status": record.source_status,
            "source_record_key": contract_source_key(record),
            "source_area_sqm": record.area_sqm,
            "source_monthly_rent_amount": record.monthly_rent_amount,
            "source_monthly_common_charge_amount": record.monthly_common_charge_amount,
            "source_monthly_parking_amount": record.monthly_parking_amount,
            "source_other_monthly_amount": record.other_monthly_amount,
            "source_monthly_total_amount": monthly_total,
            "contract_start_date": record.contract_start_date,
            "contract_end_date": record.contract_end_date,
            "raw_payload": asdict(record),
        })
    source_file_sql = "'" + workbook.name.replace("'", "''") + "'"
    payload = json.dumps(rows, ensure_ascii=False)
    sql = f"""begin;

with target_batch as (
  insert into public.rent_roll_import_batch (
    source_file_name, source_sha256, as_of_date, status, row_count, issue_count
  ) values (
    {source_file_sql}, '{source_hash}', '{as_of_date}'::date, 'uploaded', {len(rows)}, {len(issues)}
  )
  on conflict (source_sha256, as_of_date) do update set
    source_file_name = excluded.source_file_name,
    status = 'uploaded',
    row_count = excluded.row_count,
    issue_count = excluded.issue_count,
    updated_at = now()
  returning rent_roll_import_batch_id
)
insert into public.rent_roll_import_row (
  rent_roll_import_batch_id, source_sheet_name, source_row_number,
  property_name, wing_code, floor_label, unit_code, unit_type,
  tenant_code, tenant_name, source_status, source_record_key,
  source_area_sqm, source_monthly_rent_amount,
  source_monthly_common_charge_amount, source_monthly_parking_amount,
  source_other_monthly_amount,
  source_monthly_total_amount, contract_start_date, contract_end_date, raw_payload
)
select
  target_batch.rent_roll_import_batch_id,
  source.source_sheet_name, source.source_row_number,
  source.property_name, source.wing_code, source.floor_label,
  source.unit_code, source.unit_type, source.tenant_code, source.tenant_name,
  source.source_status, source.source_record_key, source.source_area_sqm,
  source.source_monthly_rent_amount, source.source_monthly_common_charge_amount,
  source.source_monthly_parking_amount, source.source_other_monthly_amount,
  source.source_monthly_total_amount,
  source.contract_start_date, source.contract_end_date, source.raw_payload
from target_batch
cross join jsonb_to_recordset($rentrollcomparison${payload}$rentrollcomparison$::jsonb) as source(
  source_sheet_name text, source_row_number integer, property_name text,
  wing_code text, floor_label text, unit_code text, unit_type text,
  tenant_code text, tenant_name text, source_status text, source_record_key text,
  source_area_sqm numeric, source_monthly_rent_amount numeric,
  source_monthly_common_charge_amount numeric, source_monthly_parking_amount numeric,
  source_other_monthly_amount numeric,
  source_monthly_total_amount numeric, contract_start_date date, contract_end_date date,
  raw_payload jsonb
)
on conflict (rent_roll_import_batch_id, source_sheet_name, source_row_number) do update set
  property_name = excluded.property_name,
  wing_code = excluded.wing_code,
  floor_label = excluded.floor_label,
  unit_code = excluded.unit_code,
  unit_type = excluded.unit_type,
  tenant_code = excluded.tenant_code,
  tenant_name = excluded.tenant_name,
  source_status = excluded.source_status,
  source_record_key = excluded.source_record_key,
  source_area_sqm = excluded.source_area_sqm,
  source_monthly_rent_amount = excluded.source_monthly_rent_amount,
  source_monthly_common_charge_amount = excluded.source_monthly_common_charge_amount,
  source_monthly_parking_amount = excluded.source_monthly_parking_amount,
  source_other_monthly_amount = excluded.source_other_monthly_amount,
  source_monthly_total_amount = excluded.source_monthly_total_amount,
  contract_start_date = excluded.contract_start_date,
  contract_end_date = excluded.contract_end_date,
  raw_payload = excluded.raw_payload,
  matched_lease_contract_unit_id = null,
  match_status = 'unmatched',
  match_note = null,
  updated_at = now();

commit;
"""
    path.write_text(sql, encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Import a rent-roll workbook into Supabase.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--report", type=Path, default=Path("rent_roll_import_report.json"))
    parser.add_argument("--apply", action="store_true", help="Write to Supabase after reviewing the report.")
    parser.add_argument("--sql-file", type=Path, help="Write a transactional SQL import for `supabase db query --linked --file`.")
    parser.add_argument("--migration-file", type=Path, help="Write an idempotent, transactional Supabase migration.")
    parser.add_argument("--comparison-sql", type=Path, help="Write SQL that stores every source row for reconciliation without changing contract data.")
    parser.add_argument("--as-of-date", help="Comparison date in YYYY-MM-DD format. Required with --comparison-sql.")
    parser.add_argument("--sheet", action="append", dest="sheet_names", help="Import only the named worksheet. Repeat to select multiple sheets.")
    args = parser.parse_args()
    records, issues = read_workbook(args.workbook, set(args.sheet_names) if args.sheet_names else None)
    report = {
        "source_file": args.workbook.name,
        "record_count": len(records),
        "issue_count": len(issues),
        "records": [asdict(record) | {"source_record_key": contract_source_key(record)} for record in records],
        "issues": [asdict(issue) for issue in issues],
    }
    args.report.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Dry-run report: {args.report} ({len(records)} records, {len(issues)} issues)")
    if not records:
        parser.error("No rent-roll records were found for the selected worksheets.")
    output_modes = [args.apply, bool(args.sql_file), bool(args.migration_file), bool(args.comparison_sql)]
    if sum(output_modes) > 1:
        parser.error("Choose exactly one of --apply, --sql-file, --migration-file, or --comparison-sql.")
    if args.comparison_sql:
        if not args.as_of_date or as_date(args.as_of_date) != args.as_of_date:
            parser.error("--comparison-sql requires --as-of-date in YYYY-MM-DD format.")
        write_reconciliation_sql(records, issues, args.workbook, args.as_of_date, args.comparison_sql)
        print(f"Prepared comparison-only SQL: {args.comparison_sql}")
        return 0
    if args.sql_file or args.migration_file:
        output_path = args.sql_file or args.migration_file
        write_sql_import(records, issues, args.workbook.name, output_path, wrap_transaction=True)
        print(f"Prepared SQL import: {output_path}")
        return 0
    if not args.apply:
        return 0
    url = os.getenv("SUPABASE_URL")
    service_role_key = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not service_role_key:
        print("SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY are required with --apply.", file=sys.stderr)
        return 2
    persisted, all_issues = persist(records, issues, args.workbook.name, SupabaseRest(url, service_role_key))
    print(f"Imported {persisted} unit rows. Recorded {len(all_issues)} review issues.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
